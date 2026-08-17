#!/usr/bin/env python3
"""Export director-friendly CfA program rosters from central Supabase data."""

from __future__ import annotations

import argparse
import json
import re
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


CFA_CLIENT_ID = "22500cd6-052a-42ff-a0cb-4f3ba9125dfd"
HEADER_FILL = PatternFill("solid", fgColor="245447")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBTLE_FILL = PatternFill("solid", fgColor="E9F1EE")


def parse_env(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    for raw_line in path.read_text().splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        values[key.strip()] = value.strip().strip('"').strip("'")
    return values


def request_json(url: str, headers: dict[str, str]) -> Any:
    request = urllib.request.Request(url, headers=headers, method="GET")
    with urllib.request.urlopen(request, timeout=120) as response:
        return json.load(response)


def supabase_list(
    base_url: str,
    table: str,
    headers: dict[str, str],
    params: dict[str, str],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    offset = 0
    while True:
        query = {**params, "limit": "1000", "offset": str(offset)}
        payload = request_json(
            f"{base_url}/rest/v1/{table}?{urllib.parse.urlencode(query)}",
            headers,
        )
        rows.extend(payload)
        if len(payload) < 1000:
            return rows
        offset += len(payload)


def parse_timestamp(value: Any) -> datetime | None:
    if not value:
        return None
    return datetime.fromisoformat(str(value).replace("Z", "+00:00"))


def access_status(enrollment: dict[str, Any], now: datetime) -> str:
    if enrollment.get("status") != "registered":
        return str(enrollment.get("status") or "unknown").title()
    if enrollment.get("revoked_at"):
        return "Revoked"
    starts = parse_timestamp(enrollment.get("access_starts_at"))
    ends = parse_timestamp(enrollment.get("access_ends_at"))
    if starts and starts > now:
        return "Scheduled"
    if ends and ends <= now:
        return "Expired"
    return "Active"


def tab_name(program: dict[str, Any]) -> str:
    name = str(program["name"])
    if name.startswith("Starlight Rays"):
        return "Starlight 26-27"
    grade = re.match(r"Grade (\d) Teaching - Renewal 2026 (In-Person|Online)", name)
    if grade:
        return f"R26 G{grade.group(1)} {grade.group(2)}"
    if name.startswith("Morning Community"):
        return "R26 Morning"
    if name.startswith("Movement Education"):
        return "R26 Movement"
    if name.startswith("Teaching Special Subjects"):
        return "R26 Special Subjects"
    cleaned = re.sub(r"[\\/*?:\[\]]", "", name)
    return cleaned[:31]


def style_table(sheet, widths: list[int]) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 24
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def main() -> None:
    dev_root = Path(__file__).resolve().parents[2]
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--supabase-env",
        type=Path,
        default=dev_root / "email-marketing-tool-1/.env",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("exports/private/cfa-program-rosters.xlsx"),
    )
    args = parser.parse_args()

    env = parse_env(args.supabase_env)
    supabase_url = (env.get("SUPABASE_URL") or env.get("VITE_SUPABASE_URL") or "").rstrip("/")
    service_key = env.get("SUPABASE_SERVICE_ROLE_KEY") or env.get("SUPABASE_SERVICE_KEY") or ""
    if not supabase_url or not service_key:
        raise RuntimeError("Supabase URL or service-role key is missing")
    headers = {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
        "Accept": "application/json",
    }

    programs = supabase_list(
        supabase_url,
        "programs",
        headers,
        {
            "client_id": f"eq.{CFA_CLIENT_ID}",
            "select": "id,name,year,format,instructor,start_date,end_date",
            "order": "name.asc",
        },
    )
    enrollments = supabase_list(
        supabase_url,
        "enrollments",
        headers,
        {
            "client_id": f"eq.{CFA_CLIENT_ID}",
            "select": (
                "id,program_id,contact_id,status,enrolled_at,source,"
                "access_starts_at,access_ends_at,revoked_at"
            ),
        },
    )
    contacts = supabase_list(
        supabase_url,
        "contacts",
        headers,
        {
            "client_id": f"eq.{CFA_CLIENT_ID}",
            "select": "id,email,first_name,last_name,company,custom_fields",
        },
    )

    programs_by_id = {row["id"]: row for row in programs}
    contacts_by_id = {row["id"]: row for row in contacts}
    roster_rows: list[dict[str, Any]] = []
    now = datetime.now(timezone.utc)
    for enrollment in enrollments:
        program = programs_by_id.get(enrollment["program_id"])
        contact = contacts_by_id.get(enrollment["contact_id"])
        if not program or not contact:
            continue
        custom_fields = contact.get("custom_fields") or {}
        roster_rows.append({
            "program_id": program["id"],
            "program": program["name"],
            "first_name": contact.get("first_name") or "",
            "last_name": contact.get("last_name") or "",
            "email": contact.get("email") or "",
            "organization": custom_fields.get("school") or contact.get("company") or "",
            "enrollment_status": str(enrollment.get("status") or "").title(),
            "access_status": access_status(enrollment, now),
            "source": enrollment.get("source") or "",
            "enrolled_at": enrollment.get("enrolled_at") or "",
            "access_starts_at": enrollment.get("access_starts_at") or "",
            "access_ends_at": enrollment.get("access_ends_at") or "",
        })
    roster_rows.sort(key=lambda row: (row["program"], row["last_name"], row["first_name"], row["email"]))

    by_program: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in roster_rows:
        by_program[row["program_id"]].append(row)
    programs_with_rosters = [program for program in programs if by_program.get(program["id"])]

    workbook = Workbook()
    workbook.remove(workbook.active)
    summary = workbook.create_sheet("Summary")
    summary.append(["CfA Program Rosters"])
    summary["A1"].font = Font(size=18, bold=True, color="245447")
    summary.append(["Generated", now.astimezone().isoformat(timespec="seconds")])
    summary.append(["Source", "Central Supabase enrollments; operational roster fields only"])
    summary.append(["Privacy", "Internal use. Payment, billing, and identity-review fields are excluded."])
    summary.append([])
    summary.append(["Program", "Year", "Format", "Instructor", "Total", "Active", "Roster tab"])
    for cell in summary[6]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    used_tabs: set[str] = {"Summary", "All Enrollments"}
    program_tabs: dict[str, str] = {}
    for program in programs_with_rosters:
        candidate = tab_name(program)
        base = candidate
        suffix = 2
        while candidate in used_tabs:
            candidate = f"{base[:27]} {suffix}"
            suffix += 1
        used_tabs.add(candidate)
        program_tabs[program["id"]] = candidate
        rows = by_program[program["id"]]
        summary.append([
            program["name"],
            program.get("year") or "",
            program.get("format") or "",
            program.get("instructor") or "",
            len(rows),
            sum(row["access_status"] == "Active" for row in rows),
            candidate,
        ])
        summary.cell(summary.max_row, 7).hyperlink = f"#'{candidate}'!A1"
    summary.freeze_panes = "A7"
    summary.auto_filter.ref = f"A6:G{summary.max_row}"
    for column, width in enumerate([58, 10, 14, 28, 10, 10, 24], start=1):
        summary.column_dimensions[get_column_letter(column)].width = width
    for row in range(7, summary.max_row + 1):
        if row % 2 == 1:
            for cell in summary[row]:
                cell.fill = SUBTLE_FILL

    headers_all = [
        "Program", "First name", "Last name", "Email", "School / organization",
        "Enrollment status", "Access status", "Source", "Enrolled at", "Access starts", "Access ends",
    ]
    all_sheet = workbook.create_sheet("All Enrollments")
    all_sheet.append(headers_all)
    for row in roster_rows:
        all_sheet.append([
            row["program"], row["first_name"], row["last_name"], row["email"], row["organization"],
            row["enrollment_status"], row["access_status"], row["source"], row["enrolled_at"],
            row["access_starts_at"], row["access_ends_at"],
        ])
    style_table(all_sheet, [58, 18, 22, 34, 32, 18, 15, 16, 24, 24, 24])

    headers_program = headers_all[1:]
    for program in programs_with_rosters:
        sheet = workbook.create_sheet(program_tabs[program["id"]])
        sheet.append(headers_program)
        for row in by_program[program["id"]]:
            sheet.append([
                row["first_name"], row["last_name"], row["email"], row["organization"],
                row["enrollment_status"], row["access_status"], row["source"], row["enrolled_at"],
                row["access_starts_at"], row["access_ends_at"],
            ])
        style_table(sheet, [18, 22, 34, 32, 18, 15, 16, 24, 24, 24])

    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    print(json.dumps({
        "output": str(args.output),
        "programs": len(programs_with_rosters),
        "enrollments": len(roster_rows),
        "active": sum(row["access_status"] == "Active" for row in roster_rows),
        "sources": dict(Counter(row["source"] for row in roster_rows)),
    }, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
